from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

ws.merge_cells('A2:D2')
ws['A2'] = 'teste de celulas'

# or equivalently
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
ws.cell(row=3, column=1, value='teste novo')
# ws.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)

wb.save('balances.xlsx')