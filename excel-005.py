from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl import load_workbook

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    if ":" in cell_range:
        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill
    else:
        cell = ws[cell_range]
        if alignment:
            cell.alignment = alignment

        cell = ws[cell_range]
        if font:
            cell.font = font

        cell.border = cell.border + top
        cell.border = cell.border + bottom

        cell.border = cell.border + left
        cell.border = cell.border + right
        if fill:
            cell.fill = fill

# wb = load_workbook("styled.xlsx")
wb = Workbook()

ws = wb.active
thin = Side(border_style="thin", color="00ff00")
double = Side(border_style="medium", color="ff00ff")
border = Border(top=double, left=thin, right=thin, bottom=double)
font = Font(b=True, color="FF0000")
fill = PatternFill("solid", fgColor="aaaaaA")
al = Alignment(horizontal="center", vertical="center")

my_cell = ws['B2']
my_cell.value = "My Cell"
style_range(ws, 'B2', border=border, fill=fill, font=font, alignment=al)

my_cell = ws['B4']
my_cell.value = "My Cell"
fill = GradientFill(stop=("000000", "FFFFFF"))
style_range(ws, 'B4:F6', border=border, fill=fill, font=font, alignment=al)

my_cell = ws['B8']
my_cell.value = "Este texto deve caber"
fill = GradientFill(stop=("000000", "FFFFFF"))
border = Border(top=double, left=thin, right=thin)
font = Font(bold=True, name="Algerian", color="FF0000")
al = Alignment(horizontal="center", vertical="center", wrap_text=True, shrinkToFit=True)
style_range(ws, 'B8:C8', border=border, fill=fill, font=font, alignment=al)
wb.save("styled.xlsx")