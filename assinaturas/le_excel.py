import openpyxl

def le_excel():
    workbook = openpyxl.load_workbook("lista.xlsx")
    sheet = workbook.active
    first_row = [] # The row where we stock the name of the column

    for col in range(2, sheet.max_column + 1):
        first_row.append(sheet.cell(row=4, column=col).value)

    data =[]
    for row in range(5, sheet.max_row + 1):
        elm = {}
        for col in range(2, sheet.max_column + 1):
            elm[first_row[col - 2]]=sheet.cell(row=row,column=col).value
        data.append(elm)

    return data
