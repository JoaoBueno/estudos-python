from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")
ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
ws3 = wb["New Title"]
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

source = wb.active
target = wb.copy_worksheet(source)

c = ws['A4']
ws['A4'] = 4
d = ws.cell(row=4, column=2, value=10)
e = ws.cell(row=5, column=2, value=10)

for i in range(1,101):
    for j in range(1,101):
        ws.cell(row=i, column=j, value=j)

cell_range = ws['A1':'C2']

wb.save('balances.xlsx')