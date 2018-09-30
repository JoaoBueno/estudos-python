#!/usr/bin/python

import sys
import csv
import string

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl import load_workbook


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    if ":" in cell_range:
        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill
    else:
        cell = ws[cell_range]
        if alignment:
            cell.alignment = alignment

        cell = ws[cell_range]
        print(cell)
        if font:
            cell.font = font

        cell.border = cell.border + top
        cell.border = cell.border + bottom

        cell.border = cell.border + left
        cell.border = cell.border + right
        if fill:
            cell.fill = fill


def processa(f):
    """
    Le arquivo e transforma em excel
    """
    wb = load_workbook("aiver135.xlsx")
    ws = wb.active
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    bordem = Border(top=medium, left=medium, right=medium, bottom=medium)
    font = Font(name='Calibri', b=True, color="000000")

    linha = csv.reader(f, delimiter=';', quotechar='|')
    for lin, row in enumerate(linha):
        if lin == 0:
            ws.cell(row=1, column=2, value=row[0].strip())
        else:
            for col, ele in enumerate(row):
                if col < 3:
                    ws.cell(row=lin + 3, column=col + 2, value=ele.strip())
                elif col == 3:
                    ws.cell(row=lin + 3, column=col +
                            2, value=float(ele.strip()))
                elif col < 6:
                    ws.cell(row=lin + 3, column=col +
                            2, value=int(ele.strip()))
                elif col == 6:
                    ws.cell(row=lin + 3, column=col + 2, value=ele.strip())
                elif col < 9:
                    ws.cell(row=lin + 3, column=col +
                            2, value=int(ele.strip()))
                else:
                    ws.cell(row=lin + 3, column=col +
                            2, value=float(ele.strip()))
                ws.cell(row=lin + 3, column=col + 2).border = border
    # linha de total
    ws.cell(row=lin + 4, column=2, value='TOTAL')
    cell_range = 'B'+str(lin+4)+':O'+str(lin+4)
    al = Alignment(horizontal="right", vertical="center")
    style_range(ws, cell_range, border=bordem, font=font, alignment=al)
    # total icms normal a recuperar
    cell_range = 'P4:P'+str(lin+3)
    ws.cell(row=lin + 4, column=16, value='=Sum('+cell_range+')')
    ws.cell(row=lin + 4, column=16).border = bordem
    # total icms/st a ressarcir
    cell_range = 'Q4:Q'+str(lin+3)
    ws.cell(row=lin + 4, column=17, value='=Sum('+cell_range+')')
    ws.cell(row=lin + 4, column=17).border = bordem

    wb.save(sys.argv[2])


if __name__ == "__main__":
    if(len(sys.argv) < 3):
        print("Use: leArq-001 arquivo arquivo-saida")
        sys.exit(0)
    try:
        with open(sys.argv[1], 'r') as f:
            processa(f)
            f.close
    except IOError:
        print(u'Arquivo não encontrado!')
        sys.exit(0)
    sys.exit(1)
